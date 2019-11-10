import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import borders, fills, colors
import os

TRUOC_PATH = "truoc"
TRUOC_COL = 9
SAU_PATH = "sau"
SAU_COL = 9
STT_STR = "STT"

def getFilesFromPath(path):
    # r=root, d=directories, f = files
    files = []
    for r, d, f in os.walk(path):
        for file in f:
            if '.xlsx' in file or '.xls' in file or '.XLS' in file or '.XLSX' in file:
                files.append(os.path.join(r, file))
    return files
  
truocFiles = getFilesFromPath(TRUOC_PATH)
sauFiles = getFilesFromPath(SAU_PATH)

nhanMap = {}
for file in sauFiles:
    wb = xlrd.open_workbook(file)
    sheet = wb.sheet_by_index(0)
    nrows = sheet.nrows
    startRow = -1
    endRow = nrows
    print(nrows)
    for i in range(nrows):
        celVal = str(sheet.cell_value(i, 0))
        celVal = celVal.strip().upper()
        print(celVal, startRow, endRow)
        if startRow >= 0 and not celVal.replace('.','',1).isdigit():
            endRow = i
            break
        if startRow == -1 and celVal == STT_STR:
            startRow = i+1
    for i in range(startRow, endRow):
        maNhan = sheet.cell_value(i, 1).strip()
        tien = sheet.cell_value(i, SAU_COL - 1)
        nhanMap[maNhan] = tien

print("danh sach ma nhan: ", nhanMap)


redpf = PatternFill(start_color=Color('FFEE1111'), end_color=Color('FFEE1111'), patternType=fills.FILL_SOLID)
yellowpf = PatternFill(start_color=Color('FFEEEE11'), end_color=Color('FFEEEE11'), patternType=fills.FILL_SOLID)

print(truocFiles)
for file in truocFiles:
    wb = xlrd.open_workbook(file)
    sheet = wb.sheet_by_index(0)
    nrows = sheet.nrows
    startRow = -1
    endRow = nrows
    nhanRows = []
    nhanThieuRows = []
    for i in range(nrows):
        celVal = str(sheet.cell_value(i, 0))
        celVal = celVal.strip().upper()
        if startRow >= 0 and not celVal.replace('.','',1).isdigit():
            endRow = i
            break
        if startRow == -1 and celVal == STT_STR:
            startRow = i+1
    for i in range(startRow, endRow):
        maNhan = str(sheet.cell_value(i, 1)).strip()
        tien = sheet.cell_value(i, TRUOC_COL - 1)
        if maNhan in nhanMap:
            if nhanMap[maNhan] == tien:
                nhanRows.append(i)
            else:
                nhanThieuRows.append(i)
    
    wb = load_workbook(filename = file)
    ws = wb.get_active_sheet()
    for r in nhanRows:
        for cell in ws[r+1]:
            cell.fill = redpf
    for r in nhanThieuRows:
        for cell in ws[r+1]:
            cell.fill = yellowpf

    wb.save(filename="output.xlsx") 